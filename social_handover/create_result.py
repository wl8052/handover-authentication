import os
import openpyxl


def create_excel_file():
    # 创建 Excel 工作簿
    wb = openpyxl.Workbook()
    ws = wb.active

    # 添加表头"user_num"和方案名称
    scheme_names = ["SREHA", "Tradition", "One_time", "Assist_program"]

    # 写入"user_num"
    ws.cell(row=1, column=1, value="user_num")

    # 写入方案名称
    for col, name in enumerate(scheme_names, start=2):
        ws.cell(row=1, column=col, value=name)

    # 保存 Excel 文件到当前执行文件旁边的目录
    current_dir = os.path.dirname(os.path.abspath(__file__))
    file_path = os.path.join(current_dir, "handover_times.xlsx")

    # 删除现有文件（如果存在）
    if os.path.exists(file_path):
        os.remove(file_path)

    wb.save(file_path)


# 调用函数创建 Excel 文件
create_excel_file()


def create_authentication_latency_excel():
    # 创建 Excel 工作簿
    wb = openpyxl.Workbook()
    ws = wb.active

    # 添加表头"user_num"和方案名称
    scheme_names = ["SREHA", "Tradition", "One_time", "Assist_program"]

    # 写入"user_num"
    ws.cell(row=1, column=1, value="user_num")

    # 写入方案名称
    for col, name in enumerate(scheme_names, start=2):
        ws.cell(row=1, column=col, value=name)

    # 保存 Excel 文件到当前执行文件旁边的目录
    current_dir = os.path.dirname(os.path.abspath(__file__))
    file_path = os.path.join(current_dir, "Authentication_latency.xlsx")

    # 删除现有文件（如果存在）
    if os.path.exists(file_path):
        os.remove(file_path)

    wb.save(file_path)


# 调用函数创建 Excel 文件
create_authentication_latency_excel()
